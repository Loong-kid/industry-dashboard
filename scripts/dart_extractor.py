# -*- coding: utf-8 -*-
"""
DART 단일판매ㆍ공급계약 체결 공시 전용 다운로더 / DB화 도구

- DART 거래소공시(pblntf_ty=I)에서 '단일판매ㆍ공급계약' 관련 공시만 골라낸다.
- 각 공시 원문(document.xml: ZIP 안 UTF-8 HTML)을 파싱해 핵심 필드를 컬럼으로 추출한다.
- 핵심 필드를 통합 CSV(계약DB.csv)에 누적(append, rcept_no 기준 중복 제거)한다.
- 원문(HTML 기본 / PDF 옵션)을 기업별 폴더에 보관한다.
- 정정공시(기재정정)는 원문 상단의 '정정관련 공시서류제출일'로 원공시와 그룹핑하여,
  그룹 내 최신 1건만 is_latest=Y, 정정 여부는 is_correction 플래그로 표시한다.

사용 예:
  # 단일 기업
  python "단일판매공급계약 추출기.py" --corp 삼성전자
  python "단일판매공급계약 추출기.py" --corp 005930 --from-year 2018

  # 관심기업 일괄 (관심기업.txt 에 한 줄에 하나씩 기업명/종목코드 기재)
  python "단일판매공급계약 추출기.py" --watchlist
  python "단일판매공급계약 추출기.py" --watchlist "C:\\경로\\관심기업.txt"

  # 원문도 PDF로 함께 보관
  python "단일판매공급계약 추출기.py" --corp 삼성전자 --doc-format both
"""
import argparse
import csv
import datetime as dt
import html
import io
import os
import re
import sys
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests

CORP_CODE_URL = "https://opendart.fss.or.kr/api/corpCode.xml"
LIST_URL = "https://opendart.fss.or.kr/api/list.json"
DOCUMENT_URL = "https://opendart.fss.or.kr/api/document.xml"
# 공개 리포이므로 키를 하드코딩하지 않는다. 환경변수 DART_API_KEY(로컬) /
# GitHub Secret DART_API_KEY(CI) 또는 --api-key 로 주입.
DEFAULT_API_KEY = ""

# 리포지토리 내부로 이식됨(구 위치: 기업\한국\_단일판매공급계약).
# 이 대시보드는 조선 4사 전용 DB를 리포 안에서 관리해 CI로 갱신한다.
BASE_DIR = Path(__file__).resolve().parent.parent / "data" / "_dart"
DB_PATH = BASE_DIR / "계약DB.csv"
DOCS_DIR = BASE_DIR / "원문"
DEFAULT_WATCHLIST_XLSX = BASE_DIR / "관심기업.xlsx"
DEFAULT_WATCHLIST_TXT = BASE_DIR / "관심기업.txt"

# 엑셀 관심기업 양식
WL_COLUMNS = ["기업명/식별자", "시작연도", "원문형식", "사용", "비고", "신규(자동)", "마지막 실행(자동)"]
DOC_FORMATS = ["html", "pdf", "both"]

# 통합 DB 컬럼 순서 (엑셀에서 바로 열림)
CSV_FIELDS = [
    "corp_name",            # 기업명
    "stock_code",           # 종목코드
    "corp_code",            # DART 고유번호
    "rcept_dt",             # 공시(접수)일자
    "deal_date",            # 계약(수주)일자
    "is_latest",            # 그룹 내 최신본 여부 (Y/'')
    "is_correction",        # 정정공시 여부 (Y/'')
    "contract_name",        # 체결계약명
    "contract_party",       # 계약상대
    "contract_amount",      # 계약금액(원)
    "recent_sales",         # 최근매출액(원)
    "sales_ratio",          # 매출액대비(%)
    "contract_start",       # 계약 시작일
    "contract_end",         # 계약 종료일
    "contract_type",        # 판매ㆍ공급계약 구분
    "region",               # 판매ㆍ공급지역
    "large_corp",           # 대규모법인여부
    "correction_target_dt", # 정정 대상(원공시) 제출일
    "correction_reason",    # 정정사유
    "report_nm",            # 보고서명
    "rcept_no",             # 접수번호
    "viewer_url",           # DART 뷰어 URL
    "file_path",            # 보관된 원문 경로
]


class DartApiError(RuntimeError):
    pass


# --------------------------------------------------------------------------- #
# 기업코드 해석
# --------------------------------------------------------------------------- #
def build_session() -> requests.Session:
    session = requests.Session()
    session.headers.update({"User-Agent": "dart-supply-contract-downloader/1.0"})
    return session


def load_corp_map(session: requests.Session, api_key: str) -> List[Dict[str, str]]:
    resp = session.get(CORP_CODE_URL, params={"crtfc_key": api_key}, timeout=30)
    resp.raise_for_status()
    with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
        with zf.open("CORPCODE.xml") as xml_file:
            root = ET.parse(xml_file).getroot()
    rows: List[Dict[str, str]] = []
    for item in root.findall("list"):
        rows.append(
            {
                "corp_code": (item.findtext("corp_code") or "").strip(),
                "corp_name": (item.findtext("corp_name") or "").strip(),
                "stock_code": (item.findtext("stock_code") or "").strip(),
            }
        )
    return rows


def resolve_corp(corp_rows: List[Dict[str, str]], query: str) -> Dict[str, str]:
    q = query.strip()
    if not q:
        raise ValueError("기업 식별자가 비어 있습니다.")

    exact_corp = [r for r in corp_rows if r["corp_code"] == q]
    exact_stock = [r for r in corp_rows if r["stock_code"] == q and r["stock_code"]]
    exact_name = [r for r in corp_rows if r["corp_name"] == q]

    candidates = exact_corp or exact_stock or exact_name
    if not candidates:
        contains = [r for r in corp_rows if q in r["corp_name"]]
        # 상장사(stock_code 보유)를 우선
        listed = [r for r in contains if r["stock_code"]]
        if len(listed) == 1:
            candidates = listed
        elif len(contains) == 1:
            candidates = contains
        elif len(contains) > 1:
            names = ", ".join(r["corp_name"] for r in contains[:10])
            raise ValueError(f"기업명이 여러 건 매칭됩니다. 더 정확히 입력하세요: {names}")

    if not candidates:
        raise ValueError(f"기업을 찾지 못했습니다: {q}")
    return candidates[0]


# --------------------------------------------------------------------------- #
# 공시 목록 수집
# --------------------------------------------------------------------------- #
def fetch_supply_contract_list(
    session: requests.Session,
    api_key: str,
    corp_code: str,
    bgn_de: str,
    end_de: str,
) -> List[Dict[str, str]]:
    results: List[Dict[str, str]] = []
    page_no = 1
    while True:
        params = {
            "crtfc_key": api_key,
            "corp_code": corp_code,
            "bgn_de": bgn_de,
            "end_de": end_de,
            "pblntf_ty": "I",  # 거래소공시 (단일판매ㆍ공급계약 포함)
            "page_count": 100,
            "page_no": page_no,
        }
        resp = session.get(LIST_URL, params=params, timeout=30)
        resp.raise_for_status()
        data = resp.json()

        status = data.get("status")
        if status == "013":  # 조회된 데이터 없음
            break
        if status != "000":
            raise DartApiError(f"list.json 오류: {status} / {data.get('message')}")

        for it in data.get("list", []):
            nm = (it.get("report_nm") or "")
            if ("단일판매" in nm) or ("공급계약" in nm):
                results.append(it)

        total_page = int(data.get("total_page", 1))
        if page_no >= total_page:
            break
        page_no += 1
    return results


# --------------------------------------------------------------------------- #
# 원문 다운로드 & 파싱
# --------------------------------------------------------------------------- #
def fetch_document_html(session: requests.Session, api_key: str, rcept_no: str) -> str:
    resp = session.get(
        DOCUMENT_URL,
        params={"crtfc_key": api_key, "rcept_no": rcept_no},
        timeout=60,
    )
    resp.raise_for_status()
    # 정상 응답은 ZIP. 오류는 XML(<result>).
    try:
        zf = zipfile.ZipFile(io.BytesIO(resp.content))
    except zipfile.BadZipFile:
        msg = resp.content[:300].decode("utf-8", errors="replace")
        raise DartApiError(f"원문 응답이 ZIP이 아닙니다: {msg}")
    raw = zf.read(zf.namelist()[0])
    return _decode_dart_html(raw)


def _decode_dart_html(raw: bytes) -> str:
    """meta 태그는 항상 euc-kr라 표기하지만 실제 바이트 인코딩은 공시 시점마다 다르다.

    최근(2023년~ 확인) 공시는 meta 표기와 무관하게 실제 UTF-8 바이트다.
    반면 2021~2022년 초의 오래된 공시는 표기대로 실제 EUC-KR/CP949 바이트라,
    UTF-8로 강제 디코드하면 한글 라벨이 통째로 U+FFFD로 치환되어 파싱이
    전멸한다(숫자는 ASCII라 살아남아 있어서 발견이 늦어지기 쉽다).
    → UTF-8 strict를 우선 시도하고 실패하면 CP949로 폴백한다.
    """
    try:
        return raw.decode("utf-8")
    except UnicodeDecodeError:
        try:
            return raw.decode("cp949")
        except UnicodeDecodeError:
            return raw.decode("utf-8", errors="replace")


def _norm(text: str) -> str:
    return re.sub(r"\s+", "", text)


def _field_of(cell: str) -> Optional[str]:
    """라벨 셀 텍스트로부터 필드명을 판정. 값 셀이면 None.

    값 셀에 '계약상대'·'계약금액' 같은 키워드가 우연히 섞여 라벨로 오인되는 것을
    막기 위해, 느슨한 부분일치가 아니라 정규화 후 정확/끝일치로 판정한다.
    (예: '계약상대방'=라벨 / '계약상대방의 영업비밀 보호요청'=값 → 구분됨)
    """
    raw = _norm(cell)
    if len(raw) > 30:                          # 긴 설명문은 라벨일 수 없음
        return None
    neg = raw.startswith("-")                  # 계약상대방 하위 항목('-최근 매출액' 등)
    core = re.sub(r"^\d+\.", "", raw.lstrip("-"))  # 선행 '-' 와 'N.' 제거

    if core in ("체결계약명", "판매ㆍ공급계약내용"):   # 양식별 계약 품목/명칭
        return "contract_name"
    if core == "판매ㆍ공급계약구분":
        return "contract_type"
    if core in ("계약상대", "계약상대방"):
        return "contract_party"
    if core == "판매ㆍ공급지역":
        return "region"
    if core == "시작일":
        return "contract_start"
    if core == "종료일":
        return "contract_end"
    if core == "대규모법인여부":
        return "large_corp"
    if core in ("매출액대비(%)", "매출액대비"):
        return "sales_ratio"
    if core == "정정사유":
        return "correction_reason"
    if "서류제출일" in core:                   # 2. 정정관련 공시서류제출일
        return "correction_target_dt"
    if re.search(r"계약\(.*?\)일자", core):    # 7. 계약(수주)일자 등
        return "deal_date"
    if core.endswith("최근매출액(원)") or core.endswith("최근매출액"):
        # 계약상대방 섹션의 '-최근 매출액(원)'(상대방 매출)은 회사 매출이 아님
        return None if neg else "recent_sales"
    if core.endswith("계약금액총액(원)") or core.endswith("계약금액(원)") or core.endswith("계약금액"):
        return "contract_amount"               # 여러 변형 중 마지막 매칭(=총액) 채택
    return None


def parse_supply_contract(doc_html: str) -> Dict[str, str]:
    """공시 원문 HTML에서 핵심 필드를 추출.

    정정공시는 상단 정정표(정정전/정정후)와 하단 본문(정정후 전체)이 함께 들어있다.
    셀을 순서대로 훑으며 라벨→다음 값 셀을 매핑하되, 같은 필드가 다시 나오면
    덮어쓴다(본문이 뒤에 오므로 최종적으로 정정후/본문 값이 남는다).
    """
    body = doc_html.split("<body>", 1)[-1]
    raw_cells = re.findall(r"<(?:td|TD)[^>]*>(.*?)</(?:td|TD)>", body, re.S)
    cells: List[str] = []
    for c in raw_cells:
        text = html.unescape(re.sub(r"\s+", " ", re.sub(r"<[^>]+>", "", c))).strip()
        if text:
            cells.append(text)

    data: Dict[str, str] = {}
    for i, cell in enumerate(cells):
        if "※" in cell and "관련공시" in cell:  # 본문 끝(※ 관련공시) 이후는 파싱 중단
            break
        fn = _field_of(cell)
        if not fn:
            continue
        nxt = cells[i + 1] if i + 1 < len(cells) else ""
        # 다음 셀이 또 다른 라벨이면 값이 비어있는 것으로 간주 (덮어쓰지 않음)
        if nxt and _field_of(nxt) is None:
            data[fn] = nxt

    # 금액류는 콤마/공백 제거하여 순수 숫자로 (엑셀 수치 처리 용이)
    for key in ("contract_amount", "recent_sales"):
        if data.get(key):
            digits = re.sub(r"[^\d]", "", data[key])
            if digits:
                data[key] = digits
    return data


def _cell_text(html_fragment: str) -> str:
    return html.unescape(re.sub(r"\s+", " ", re.sub(r"<[^>]+>", "", html_fragment))).strip()


def _label_key(text: str) -> str:
    """동적 컬럼 키 정규화: 선행 하이픈 제거 + 공백 제거.
    양식별 표기차('최근 매출액' vs '최근매출액', '- 주요사업' vs '-주요사업')를 한 컬럼으로 통일."""
    return re.sub(r"\s+", "", text.lstrip("-").strip())


def parse_all_fields(doc_html: str) -> Dict[str, str]:
    """원문 표의 모든 항목을 빠짐없이 추출. 양식별 부가 항목까지 보존한다.

    표 행(<tr>) 구조:
      - 2칸: (라벨, 값)
      - 3칸: (섹션헤더, 하위라벨, 값)  ← 섹션 첫 줄에 하위항목 포함
      - 1칸: 섹션헤더 단독 또는 설명문
    같은 라벨이 섹션마다 또 나오므로(예: '최근 매출액(원)'이 계약내역/계약상대방 양쪽에)
    키에 섹션명을 prefix 하여 충돌을 막는다. 반환 키는 한글 그대로(엑셀 컬럼명).
    정정공시는 본문('1. 판매ㆍ공급계약…')부터 파싱하여 상단 정정표 노이즈를 건너뛴다.
    """
    body = doc_html.split("<body>", 1)[-1]
    grid: List[List[str]] = []
    for tr in re.findall(r"<tr[^>]*>(.*?)</tr>", body, re.S | re.I):
        tds = re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", tr, re.S | re.I)
        cells = [c for c in (_cell_text(t) for t in tds) if c != ""]
        if cells:
            grid.append(cells)

    # 본문 시작 지점(정정표 건너뛰기): '1.' + '판매'·'공급계약' 포함 섹션 행
    start = 0
    for i, cells in enumerate(grid):
        if re.match(r"^\s*1\.", cells[0]) and "판매" in cells[0] and "공급계약" in cells[0]:
            start = i
            break

    out: Dict[str, str] = {}
    section = ""
    for cells in grid[start:]:
        if any(("※" in c and "관련공시" in c) for c in cells):
            break
        first = cells[0]
        if re.match(r"^\s*\d+\.", first):            # 섹션 헤더 행
            section = _label_key(re.sub(r"^\s*\d+\.\s*", "", first))
            if len(cells) == 2:                      # 섹션 자체가 값 보유 (계약상대방 등)
                out[section] = cells[1]
            elif len(cells) >= 3:                    # 섹션 첫 줄에 하위항목
                out[f"{section}_{_label_key(cells[1])}"] = cells[2]
        else:                                        # 하위 항목 행
            if len(cells) >= 2:
                key = f"{section}_{_label_key(cells[0])}" if section else _label_key(cells[0])
                out[key] = cells[-1]
            elif len(cells) == 1:                    # 설명문 누적
                key = section or "비고"
                out[key] = (out.get(key, "") + (" " if out.get(key) else "") + cells[0])
    return out


# --- 원문 PDF (옵션) --------------------------------------------------------- #
def get_dcm_no(session: requests.Session, rcept_no: str) -> Optional[str]:
    resp = session.get(
        f"https://dart.fss.or.kr/dsaf001/main.do?rcpNo={rcept_no}", timeout=30
    )
    resp.raise_for_status()
    m = re.search(r"node[12]?\['dcmNo'\]\s*=\s*\"(\d+)\"", resp.text)
    return m.group(1) if m else None


def download_pdf_bytes(session: requests.Session, rcept_no: str, dcm_no: str) -> bytes:
    referer = f"https://dart.fss.or.kr/pdf/download/main.do?rcp_no={rcept_no}&dcm_no={dcm_no}"
    session.get(referer, timeout=30).raise_for_status()
    resp = session.get(
        f"https://dart.fss.or.kr/pdf/download/pdf.do?rcp_no={rcept_no}&dcm_no={dcm_no}",
        headers={"Referer": referer, "User-Agent": "Mozilla/5.0"},
        timeout=60,
    )
    resp.raise_for_status()
    if "application/pdf" not in (resp.headers.get("Content-Type") or "").lower() or not resp.content:
        raise DartApiError("PDF 다운로드 응답이 비정상입니다.")
    return resp.content


# --------------------------------------------------------------------------- #
# 통합 DB
# --------------------------------------------------------------------------- #
def safe_name(name: str) -> str:
    return re.sub(r"[\\/:*?\"<>|]+", "_", name).strip()


def atomic_save_workbook(wb, path: Path) -> None:
    """임시 파일에 저장 후 원본으로 교체(원자적). 저장 중 강제 종료돼도 원본이 안 깨진다."""
    tmp = path.with_name(path.name + ".tmp")
    wb.save(tmp)
    tmp.replace(path)


def load_existing_db() -> Dict[str, Dict[str, str]]:
    """rcept_no -> 행 dict"""
    rows: Dict[str, Dict[str, str]] = {}
    if DB_PATH.exists():
        with DB_PATH.open("r", encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f):
                rid = row.get("rcept_no")
                if rid:
                    rows[rid] = row
    return rows


def recompute_latest(rows: List[Dict[str, str]]) -> None:
    """정정 그룹핑: (corp_code, group_key) 단위로 최신 1건만 is_latest=Y.

    같은 계약의 원공시·정정공시들은 계약(수주)일자(deal_date)가 동일하게 유지되므로
    이를 1차 키로 사용한다(양식에 따라 '정정관련 공시서류제출일'이 정정본 자신의
    날짜를 가리키는 경우가 있어 신뢰할 수 없음). deal_date가 없으면 계약기간으로,
    그것도 없으면 단독 건으로 본다."""
    for r in rows:
        corp = r.get("corp_code", "")
        deal = re.sub(r"[^\d]", "", r.get("deal_date", ""))
        if deal:
            r["_gkey"] = (corp, "D", deal)
        elif r.get("contract_start") or r.get("contract_end"):
            r["_gkey"] = (corp, "P", r.get("contract_start", ""), r.get("contract_end", ""))
        else:
            r["_gkey"] = (corp, "R", r.get("rcept_no", ""))

    latest_no: Dict[Tuple, str] = {}
    for r in rows:
        gk = r["_gkey"]
        cur = (r.get("rcept_dt", ""), r.get("rcept_no", ""))
        if gk not in latest_no or cur > latest_no[gk][0]:
            latest_no[gk] = (cur, r.get("rcept_no", ""))

    winners = {v[1] for v in latest_no.values()}
    for r in rows:
        r["is_latest"] = "Y" if r.get("rcept_no") in winners else ""
        r.pop("_gkey", None)


def write_db(rows: List[Dict[str, str]]) -> None:
    BASE_DIR.mkdir(parents=True, exist_ok=True)
    # 정렬: 기업명 → 공시일 내림차순
    rows.sort(key=lambda r: (r.get("corp_name", ""), r.get("rcept_dt", "")), reverse=False)
    rows.sort(key=lambda r: r.get("rcept_dt", ""), reverse=True)
    rows.sort(key=lambda r: r.get("corp_name", ""))

    # 핵심 컬럼 뒤에 양식별 부가 컬럼을 등장 순서대로 합집합
    extra: List[str] = []
    seen = set(CSV_FIELDS)
    for r in rows:
        for k in r:
            if k not in seen and not k.startswith("_"):
                seen.add(k)
                extra.append(k)
    fields = CSV_FIELDS + extra

    # 원자적 저장: 임시 파일에 쓰고 교체 (저장 중 강제 종료돼도 기존 DB 보존)
    tmp = DB_PATH.with_name(DB_PATH.name + ".tmp")
    with tmp.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    tmp.replace(DB_PATH)


# --------------------------------------------------------------------------- #
# 한 기업 처리
# --------------------------------------------------------------------------- #
def process_corp(
    session: requests.Session,
    api_key: str,
    corp: Dict[str, str],
    bgn_de: str,
    end_de: str,
    db: Dict[str, Dict[str, str]],
    doc_format: str,
    refresh: bool = False,
    no_docs: bool = False,
) -> int:
    corp_code = corp["corp_code"]
    corp_name = corp["corp_name"]
    stock_code = corp.get("stock_code", "")

    listed = fetch_supply_contract_list(session, api_key, corp_code, bgn_de, end_de)
    if not listed:
        print(f"  [{corp_name}] 단일판매ㆍ공급계약 공시 없음")
        return 0

    corp_doc_dir = DOCS_DIR / safe_name(corp_name)
    new_count = 0

    for it in listed:
        rcept_no = it.get("rcept_no", "")
        report_nm = (it.get("report_nm") or "").strip()
        rcept_dt = it.get("rcept_dt", "")
        existing = db.get(rcept_no)
        # 스킵 판단. no_docs(CI)는 원문 파일을 남기지 않으므로 DB 존재만으로 스킵.
        # 로컬(원문 보관)은 원문 파일까지 있어야 스킵(파일 유실 시 재수집).
        if not refresh and existing:
            if no_docs or (existing.get("file_path") and Path(existing["file_path"]).exists()):
                continue

        try:
            doc_html = fetch_document_html(session, api_key, rcept_no)
        except Exception as ex:
            print(f"  [WARN] {rcept_no} 원문 수집 실패: {ex}")
            continue

        parsed = parse_supply_contract(doc_html)
        all_fields = parse_all_fields(doc_html)  # 양식별 부가 항목 전부
        is_correction = "Y" if ("정정" in report_nm or parsed.get("correction_target_dt")) else ""

        # 원문 보관 (no_docs 모드는 파싱만 하고 파일은 남기지 않음 → CI/증분용)
        file_path = ""
        if not no_docs:
            corp_doc_dir.mkdir(parents=True, exist_ok=True)
        if not no_docs and doc_format in ("html", "both"):
            html_name = f"{safe_name(corp_name)}_{rcept_dt}_{rcept_no}.html"
            html_path = corp_doc_dir / html_name
            # 브라우저에서 바로 열리도록 meta charset을 utf-8로 교정
            fixed = re.sub(r"charset\s*=\s*euc-kr", "charset=utf-8", doc_html, flags=re.I)
            html_path.write_text(fixed, encoding="utf-8")
            file_path = str(html_path.resolve())
        if not no_docs and doc_format in ("pdf", "both"):
            try:
                dcm = get_dcm_no(session, rcept_no)
                if dcm:
                    pdf_bytes = download_pdf_bytes(session, rcept_no, dcm)
                    pdf_name = f"{safe_name(corp_name)}_{rcept_dt}_{rcept_no}.pdf"
                    pdf_path = corp_doc_dir / pdf_name
                    pdf_path.write_bytes(pdf_bytes)
                    if not file_path:
                        file_path = str(pdf_path.resolve())
            except Exception as ex:
                print(f"  [WARN] {rcept_no} PDF 다운로드 실패: {ex}")

        row = {
            "corp_name": corp_name,
            "stock_code": stock_code,
            "corp_code": corp_code,
            "rcept_dt": rcept_dt,
            "deal_date": parsed.get("deal_date", ""),
            "is_latest": "",  # 나중에 일괄 재계산
            "is_correction": is_correction,
            "contract_name": parsed.get("contract_name", ""),
            "contract_party": parsed.get("contract_party", ""),
            "contract_amount": parsed.get("contract_amount", ""),
            "recent_sales": parsed.get("recent_sales", ""),
            "sales_ratio": parsed.get("sales_ratio", ""),
            "contract_start": parsed.get("contract_start", ""),
            "contract_end": parsed.get("contract_end", ""),
            "contract_type": parsed.get("contract_type", ""),
            "region": parsed.get("region", ""),
            "large_corp": parsed.get("large_corp", ""),
            "correction_target_dt": parsed.get("correction_target_dt", ""),
            "correction_reason": parsed.get("correction_reason", ""),
            "report_nm": report_nm,
            "rcept_no": rcept_no,
            "viewer_url": f"https://dart.fss.or.kr/dsaf001/main.do?rcpNo={rcept_no}",
            "file_path": file_path,
        }
        # 양식별 부가 항목을 동적 컬럼으로 덧붙임 (핵심 컬럼명과 충돌 없음: 한글 키)
        for k, v in all_fields.items():
            if k not in row:
                row[k] = v
        db[rcept_no] = row
        new_count += 1
        amt = row["contract_amount"]
        amt_disp = f"{int(amt):,}" if amt.isdigit() else (amt or "-")
        flag = " [정정]" if is_correction else ""
        print(f"  [OK]{flag} {rcept_dt} {report_nm} / 계약상대={row['contract_party'] or '-'} / 금액={amt_disp} / 매출대비={row['sales_ratio'] or '-'}%")

    print(f"  [{corp_name}] 신규 {new_count}건 / 조회 {len(listed)}건")
    return new_count


# --------------------------------------------------------------------------- #
# watchlist
# --------------------------------------------------------------------------- #
def ensure_watchlist_template(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "# 단일판매ㆍ공급계약 관심기업 목록\n"
        "# 한 줄에 하나씩: 기업명 또는 종목코드(6자리) 또는 DART 고유번호(8자리)\n"
        "# '#' 으로 시작하는 줄은 주석입니다.\n"
        "\n"
        "삼성전자\n"
        "# 005930\n",
        encoding="utf-8",
    )


def read_watchlist(path: Path) -> List[str]:
    items: List[str] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        s = line.strip()
        if s and not s.startswith("#"):
            items.append(s)
    return items


# --------------------------------------------------------------------------- #
# 엑셀 관심기업 양식
# --------------------------------------------------------------------------- #
def create_watchlist_xlsx(path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "관심기업"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    guide_fill = PatternFill("solid", fgColor="FFF2CC")
    example_fill = PatternFill("solid", fgColor="E8F1FA")
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="top")

    for col_idx, name in enumerate(WL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # 가이드 행 (사용=가이드 → 자동 스킵)
    guide_row = [
        "기업명(삼성전자) / 종목코드(005930) / 고유번호(8자리)",
        "정수, 비우면 5년 전부터",
        "html / pdf / both  (비우면 html)",
        "가이드",
        "메모 자유 입력",
        "실행하면 새로 받은 건수 자동 기록",
        "실행 시각 자동 기록",
    ]
    for col_idx, val in enumerate(guide_row, start=1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.fill = guide_fill
        cell.alignment = wrap
    ws.row_dimensions[2].height = 45

    examples = [
        ["삼성전자", None, "html", "Y", "", "", ""],
        ["라온텍", 2023, "html", "Y", "예: 시작연도 지정", "", ""],
    ]
    for row_idx, row in enumerate(examples, start=3):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val).fill = example_fill

    dv_fmt = DataValidation(type="list", formula1=f'"{",".join(DOC_FORMATS)}"', allow_blank=True)
    dv_use = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(dv_fmt)
    ws.add_data_validation(dv_use)
    dv_fmt.add("C3:C1000")
    dv_use.add("D3:D1000")

    widths = {"A": 30, "B": 10, "C": 12, "D": 8, "E": 24, "F": 12, "G": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"
    atomic_save_workbook(wb, path)


def read_watchlist_xlsx(path: Path) -> Tuple[object, object, List[Dict]]:
    """엑셀을 열어 (workbook, worksheet, 처리할 행 목록) 반환.
    각 행: {row, ident, from_year(Optional[int]), doc_format(Optional[str])}"""
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb.active
    rows: List[Dict] = []
    for r in range(2, ws.max_row + 1):
        ident = ws.cell(row=r, column=1).value
        ident = (str(ident).strip() if ident is not None else "")
        use = ws.cell(row=r, column=4).value
        use = (str(use).strip() if use is not None else "")
        if not ident:
            continue
        if use and use.upper() not in ("Y", "예", "O"):  # N/가이드 등은 스킵
            continue

        yr_raw = ws.cell(row=r, column=2).value
        try:
            from_year = int(yr_raw) if yr_raw not in (None, "") else None
        except (TypeError, ValueError):
            from_year = None

        fmt = ws.cell(row=r, column=3).value
        fmt = (str(fmt).strip().lower() if fmt else "")
        if fmt not in DOC_FORMATS:
            fmt = None

        rows.append({"row": r, "ident": ident, "from_year": from_year, "doc_format": fmt})
    return wb, ws, rows


# --------------------------------------------------------------------------- #
def _default_bgn_end(from_year: Optional[int]) -> Tuple[str, str]:
    today = dt.date.today()
    fy = from_year if from_year else (today.year - 5)
    return f"{fy}0101", today.strftime("%Y%m%d")


def run_excel_mode(args, wl_path: Path) -> int:
    """엑셀 관심기업 양식을 읽어 일괄 처리하고 결과를 엑셀에 기록."""
    try:
        wb, ws, rows = read_watchlist_xlsx(wl_path)
    except ImportError:
        print("openpyxl 이 필요합니다. pip install openpyxl", file=sys.stderr)
        return 2

    if not rows:
        print(f"처리할 기업이 없습니다. 엑셀의 '사용' 칸과 '기업명/식별자' 칸을 확인하세요:\n  {wl_path.resolve()}")
        return 1

    session = build_session()
    print("기업코드 매핑 로딩 중...")
    corp_rows = load_corp_map(session, args.api_key)
    db = load_existing_db()

    end_de = dt.date.today().strftime("%Y%m%d")
    total_new = 0
    for item in rows:
        try:
            corp = resolve_corp(corp_rows, item["ident"])
        except ValueError as ex:
            print(f"[SKIP] {item['ident']}: {ex}")
            ws.cell(row=item["row"], column=6, value="식별실패")
            ws.cell(row=item["row"], column=7, value=dt.datetime.now().strftime("%Y-%m-%d %H:%M"))
            atomic_save_workbook(wb, wl_path)
            continue

        from_year = item["from_year"] if item["from_year"] else args.from_year
        bgn_de, _ = _default_bgn_end(from_year)
        doc_format = item["doc_format"] or args.doc_format

        print(f"\n▶ {corp['corp_name']} ({corp['corp_code']}, {corp.get('stock_code') or '비상장'})")
        n = process_corp(session, args.api_key, corp, bgn_de, end_de, db, doc_format, args.refresh)
        total_new += n

        ws.cell(row=item["row"], column=6, value=n)
        ws.cell(row=item["row"], column=7, value=dt.datetime.now().strftime("%Y-%m-%d %H:%M"))
        atomic_save_workbook(wb, wl_path)  # 행마다 저장 (중단 대비)

    rows_all = list(db.values())
    recompute_latest(rows_all)
    write_db(rows_all)

    print("\n완료")
    print(f"신규 수집: {total_new}건 / 전체 DB: {len(rows_all)}건")
    print(f"관심기업 엑셀: {wl_path.resolve()}")
    print(f"통합 DB: {DB_PATH.resolve()}")
    print(f"원문 폴더: {DOCS_DIR.resolve()}")
    return 0


def run_query_mode(args, queries: List[str]) -> int:
    """--corp 단일 또는 .txt 목록 일괄 처리."""
    session = build_session()
    print("기업코드 매핑 로딩 중...")
    corp_rows = load_corp_map(session, args.api_key)
    db = load_existing_db()

    bgn_de, end_de = _default_bgn_end(args.from_year)
    total_new = 0
    for q in queries:
        try:
            corp = resolve_corp(corp_rows, q)
        except ValueError as ex:
            print(f"[SKIP] {q}: {ex}")
            continue
        print(f"\n▶ {corp['corp_name']} ({corp['corp_code']}, {corp.get('stock_code') or '비상장'})")
        total_new += process_corp(session, args.api_key, corp, bgn_de, end_de, db,
                                  args.doc_format, args.refresh, getattr(args, "no_docs", False))

    rows_all = list(db.values())
    recompute_latest(rows_all)
    write_db(rows_all)

    print("\n완료")
    print(f"신규 수집: {total_new}건 / 전체 DB: {len(rows_all)}건")
    print(f"통합 DB: {DB_PATH.resolve()}")
    print(f"원문 폴더: {DOCS_DIR.resolve()}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(
        description="DART 단일판매ㆍ공급계약 공시 전용 다운로더/DB화"
    )
    parser.add_argument("--corp", help="기업명 또는 종목코드(6자리) 또는 고유번호(8자리)")
    parser.add_argument(
        "--watchlist",
        nargs="?",
        const="__default__",
        help="관심기업 목록 일괄 처리. 경로 생략 시 기본 관심기업.xlsx (.txt 도 지원)",
    )
    parser.add_argument(
        "--api-key",
        default=os.getenv("DART_API_KEY", DEFAULT_API_KEY),
        help="OpenDART API KEY (미입력 시 기본 키)",
    )
    parser.add_argument(
        "--from-year", type=int, default=None, help="검색 시작 연도(기본: 현재연도-5)"
    )
    parser.add_argument(
        "--doc-format",
        choices=["html", "pdf", "both"],
        default="html",
        help="원문 보관 형식 기본값 (엑셀 행값이 우선, 기본: html)",
    )
    parser.add_argument(
        "--refresh",
        action="store_true",
        help="이미 받은 공시도 원문을 다시 파싱해 DB 갱신(부가 컬럼 채우기 등 스키마 변경 시)",
    )
    parser.add_argument(
        "--no-docs",
        action="store_true",
        help="원문 파일을 저장하지 않고 파싱만(증분은 DB 존재로만 판단). CI/서버용",
    )
    args = parser.parse_args()

    if not args.api_key:
        print("오류: API 키가 없습니다.", file=sys.stderr)
        return 2

    # 1) 단일 기업
    if args.corp:
        return run_query_mode(args, [args.corp])

    # 2) watchlist 경로 결정 (.xlsx 우선)
    if args.watchlist and args.watchlist != "__default__":
        wl_path = Path(args.watchlist)
    elif DEFAULT_WATCHLIST_XLSX.exists():
        wl_path = DEFAULT_WATCHLIST_XLSX
    elif DEFAULT_WATCHLIST_TXT.exists():
        wl_path = DEFAULT_WATCHLIST_TXT
    else:
        wl_path = DEFAULT_WATCHLIST_XLSX  # 기본 엑셀을 새로 생성

    # 3) 파일 없으면 양식 생성 후 안내
    if not wl_path.exists():
        if wl_path.suffix.lower() == ".xlsx":
            try:
                create_watchlist_xlsx(wl_path)
            except ImportError:
                print("openpyxl 이 필요합니다. pip install openpyxl", file=sys.stderr)
                return 2
        else:
            ensure_watchlist_template(wl_path)
        print(f"관심기업 양식을 새로 만들었습니다:\n  {wl_path.resolve()}")
        print("엑셀에 기업을 적은 뒤 다시 실행하거나, --corp 로 단일 기업을 지정하세요.")
        return 0

    # 4) 확장자에 따라 처리
    if wl_path.suffix.lower() == ".xlsx":
        return run_excel_mode(args, wl_path)
    queries = read_watchlist(wl_path)
    if not queries:
        print("처리할 기업이 없습니다. 관심기업 파일을 확인하세요.")
        return 1
    return run_query_mode(args, queries)


if __name__ == "__main__":
    raise SystemExit(main())
