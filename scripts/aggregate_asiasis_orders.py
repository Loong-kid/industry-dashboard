# -*- coding: utf-8 -*-
"""asiasis 신조프로젝트 오더북(정규화 Excel) → data/shipbuilding/asiasis_orders.json.

로컬 전용(CI 아님) — extract_shinyoung.py 와 같은 성격. 원본 데이터셋(일간조선해양
크롤)은 별도 비공개 폴더(../asiasis-orderbook)에 있고, 여기서는 그 정규화 산출물을
대시보드가 읽는 테이블 JSON으로 변환만 한다.

기본 소스: ../asiasis-orderbook/신조프로젝트_정규화.xlsx (normalize.py 산출물)
    python scripts/aggregate_asiasis_orders.py
    python scripts/aggregate_asiasis_orders.py --src <다른경로.xlsx>
"""
import argparse
import json
import re
from collections import Counter
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
REPO = HERE.parent
DEFAULT_SRC = REPO.parent / "asiasis-orderbook" / "신조프로젝트_정규화.xlsx"
OUT = REPO / "data" / "shipbuilding" / "asiasis_orders.json"

# normalize.py 의 COLUMNS 순서와 1:1 대응
COL = {name: i for i, name in enumerate(
    ["번호", "보고일", "제목", "선종", "선종(원문)", "사이즈", "납기", "납기연도",
     "조선소", "국적", "조선소(원문)", "발주처", "척수", "선가($m)", "선가기준",
     "선가(원문)", "Remarks", "bbs_no", "URL"])}


def s(v):
    v = ("" if v is None else str(v)).strip()
    return "" if v in ("-", "nan", "None") else v


# normalize.py VESSEL_RULES 의 canonical 결과값 = 필터 칩으로 쓸 굵은 분류.
# 이 집합에 없는(미분류·원문유지) 선종은 전부 "기타"로 버킷팅해 칩 수 폭발 방지.
CANON_CATEGORIES = {
    "VLCC(초대형원유운반선)", "원유운반선", "석유제품/케미컬선", "셔틀탱커", "탱커(기타)",
    "LNG운반선", "가스선(LPG/암모니아/에탄)", "컨테이너선", "벌커", "자동차운반선(PCTC)",
    "크루즈/여객선", "해양설비", "해상풍력설치선", "군함/방산",
}


def category_of(vessel_type):
    return vessel_type if vessel_type in CANON_CATEGORIES else "기타"


def run(src: Path):
    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb.active
    rit = ws.iter_rows(values_only=True)
    next(rit)  # 헤더 스킵

    orders = []
    for r in rit:
        report_date = s(r[COL["보고일"]])
        if not re.match(r"\d{4}-\d{2}-\d{2}", report_date):
            continue  # 날짜 정규화 실패 행은 테이블에서 제외
        count = r[COL["척수"]]
        price = r[COL["선가($m)"]]
        vtype = s(r[COL["선종"]])
        orders.append({
            "report_date": report_date,
            "title": s(r[COL["제목"]]),
            "vessel_type": vtype,
            "category": category_of(vtype),
            "vessel_type_raw": s(r[COL["선종(원문)"]]),
            "size": s(r[COL["사이즈"]]),
            "delivery": s(r[COL["납기"]]),
            "delivery_year": s(r[COL["납기연도"]]),
            "builder": s(r[COL["조선소"]]),
            "nationality": s(r[COL["국적"]]) or "미상",
            "buyer": s(r[COL["발주처"]]),
            "count": int(count) if isinstance(count, (int, float)) else None,
            "price_m": float(price) if isinstance(price, (int, float)) else None,
            "price_basis": s(r[COL["선가기준"]]),
            "price_raw": s(r[COL["선가(원문)"]]),
            "url": s(r[COL["URL"]]),
        })

    orders.sort(key=lambda o: o["report_date"], reverse=True)

    # 칩 목록: 빈도 높은 순. "기타"는 항상 맨 끝으로.
    nat_freq = Counter(o["nationality"] for o in orders)
    cat_freq = Counter(o["category"] for o in orders if o["category"])
    nationalities = [n for n, _ in nat_freq.most_common()]
    categories = [c for c, _ in cat_freq.most_common() if c != "기타"]
    if any(o["category"] == "기타" for o in orders):
        categories.append("기타")

    doc = {
        "id": "asiasis_orders",
        "name": "글로벌 신조프로젝트 오더북",
        "unit": "$m/vessel",
        "frequency": "수시(보고 발생 시)",
        "source": "일간조선해양 (asiasis.com)",
        "source_url": "http://asiasis.com/wi_bbs/wi_kr_list.php?bbs_arr=1",
        "note": "전세계 신조 발주 프로젝트. 국내 4사 DART 수주와 별개 소스로, 교차검증·디커플링 확인용.",
        "updated": orders[0]["report_date"] if orders else date.today().isoformat(),
        "nationalities": nationalities,
        "categories": categories,
        "orders": orders,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(doc, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"완료: {len(orders):,}건 → {OUT}")
    print(f"  국적 {len(nationalities)}종, 선종 {len(categories)}종, "
          f"기간 {orders[-1]['report_date']}~{orders[0]['report_date']}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", type=Path, default=DEFAULT_SRC)
    args = ap.parse_args()
    if not args.src.exists():
        raise SystemExit(f"소스 없음: {args.src}\n  먼저 asiasis-orderbook/normalize.py 실행 필요")
    run(args.src)
