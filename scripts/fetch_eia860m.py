# -*- coding: utf-8 -*-
"""EIA-860M (Preliminary Monthly Electric Generator Inventory) 수집기.

미국 1MW 이상 전 발전기 명부. 매월 EIA가 xlsx 한 개로 발행한다.
시트: Operating / Planned / Retired / Canceled or Postponed (+ *_PR 푸에르토리코).

이 스크립트가 만드는 것 (data/_eia/):
  vintages.csv.gz   빈티지 저장소. 매 발행월(vintage)의 Planned 시트 전 행.
                    → 월간 diff로 COD 지연 / 신규진입 / 단계승급 / 이탈 추적. 이게 핵심 자산.
  dim.csv.gz        (plant_id, gen_id) 차원표. 이름·위치·BA 등 최신값.
  operating.csv.gz  최신 스냅샷의 Operating 시트(가동중 전 발전기 + 은퇴예정일).
  retired.csv.gz    최신 스냅샷의 Retired 시트.
  canceled.csv.gz   최신 스냅샷의 Canceled or Postponed 시트.

사용법:
  python scripts/fetch_eia860m.py --backfill          # 최초 1회(로컬). 2021-01부터 전부.
  python scripts/fetch_eia860m.py                     # CI 기본. 신규 발행월만 증분.

함정 정리:
  1) 최신월 파일은 `xls/`, 과거월은 `archive/xls/`에 있다. 반대쪽 경로도 **404가 아니라
     HTTP 200 + 67KB짜리 stub**을 돌려준다. 상태코드로 판단하면 그대로 속는다.
     → 시트 구성으로 검증(_open_validated).
  2) **컬럼 순서가 시대별로 다르다.** 2021년판은 Sector가 5번째, 2023년판 이후는 Google Map이
     5번째다. 이름은 같으므로 반드시 헤더명으로 읽어야 한다(위치 인덱싱 금지).
  3) 헤더는 3행(0-based 2행). 위 2행은 제목/공백.
  4) 빈 셀이 NaN이 아니라 **공백 문자 한 칸(' ')**이다. 숫자 파싱 시 그대로 터진다.
  5) Nameplate Energy Capacity (MWh) / DC Net Capacity 컬럼은 **2022년 이후 판에만** 있다.
     없는 빈티지는 결측으로 둔다.
  6) Planned 시트에는 MWh·DC 컬럼이 아예 없다(준공 후 Operating에서만 붙음).
  7) 푸에르토리코는 별도 시트(*_PR)다. 본토와 섞지 않는다(수집도 하지 않음).
"""
import argparse
import gzip
import io
import os
import shutil
import sys
import time
import urllib.error
import urllib.request
from concurrent.futures import ThreadPoolExecutor
from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
EIA_DIR = ROOT / "data" / "_eia"
RAW_DIR = EIA_DIR / "raw"  # gitignore 대상

BASE = "https://www.eia.gov/electricity/data/eia860m"
MONTHS = ["january", "february", "march", "april", "may", "june",
          "july", "august", "september", "october", "november", "december"]
UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}

BACKFILL_START = (2021, 1)
EXPECTED_SHEETS = {"Operating", "Planned", "Retired"}
STUB_MAX_BYTES = 500_000  # 정상 파일은 8MB+. stub은 67KB.

# 시트별로 뽑을 컬럼: {표준명: 엑셀 헤더명}
PLANNED_COLS = {
    "plant_id": "Plant ID", "gen_id": "Generator ID", "status": "Status",
    "tech": "Technology", "mw": "Nameplate Capacity (MW)",
    "cod_y": "Planned Operation Year", "cod_m": "Planned Operation Month",
}
DIM_COLS = {
    "plant_id": "Plant ID", "gen_id": "Generator ID", "entity_id": "Entity ID",
    "entity": "Entity Name", "plant": "Plant Name", "state": "Plant State",
    "county": "County", "ba": "Balancing Authority Code", "sector": "Sector",
    "tech": "Technology", "energy_src": "Energy Source Code",
    "prime_mover": "Prime Mover Code", "lat": "Latitude", "lon": "Longitude",
}
OPERATING_COLS = {
    **DIM_COLS,
    "mw": "Nameplate Capacity (MW)", "mw_summer": "Net Summer Capacity (MW)",
    "mwh": "Nameplate Energy Capacity (MWh)", "dc_mw": "DC Net Capacity (MW)",
    "op_y": "Operating Year", "op_m": "Operating Month", "status": "Status",
    "ret_y": "Planned Retirement Year", "ret_m": "Planned Retirement Month",
}
RETIRED_COLS = {
    **DIM_COLS,
    "mw": "Nameplate Capacity (MW)", "mw_summer": "Net Summer Capacity (MW)",
    "op_y": "Operating Year", "op_m": "Operating Month",
    "ret_y": "Retirement Year", "ret_m": "Retirement Month",
}
CANCELED_COLS = {
    **DIM_COLS,
    "mw": "Nameplate Capacity (MW)", "mw_summer": "Net Summer Capacity (MW)",
}


# ---------------------------------------------------------------- 다운로드

def _month_iter(start, end):
    y, m = start
    while (y, m) <= end:
        yield y, m
        m += 1
        if m > 12:
            y, m = y + 1, 1


def _download(url: str, dest: Path) -> bool:
    req = urllib.request.Request(url, headers=UA)
    try:
        with urllib.request.urlopen(req, timeout=180) as r:
            data = r.read()
    except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError):
        return False
    if len(data) < STUB_MAX_BYTES:
        return False  # 함정 1: stub
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(data)
    return True


def fetch_month(y: int, m: int, force: bool = False) -> Path | None:
    """해당 발행월 xlsx를 raw 캐시에 확보하고 경로 반환. 없으면 None."""
    dest = RAW_DIR / f"{y:04d}-{m:02d}.xlsx"
    if dest.exists() and dest.stat().st_size > STUB_MAX_BYTES and not force:
        return dest
    name = f"{MONTHS[m - 1]}_generator{y}.xlsx"
    # 함정 1: 최신월은 xls/, 과거월은 archive/xls/. 둘 다 시도.
    for sub in ("archive/xls", "xls"):
        if _download(f"{BASE}/{sub}/{name}", dest):
            return dest
    return None


# ---------------------------------------------------------------- 파싱

def _open_validated(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if not EXPECTED_SHEETS.issubset(set(wb.sheetnames)):
        wb.close()
        raise ValueError(f"예상 시트 없음 (stub 의심): {path.name} -> {wb.sheetnames}")
    return wb


def _blank(v):
    """함정 4: 빈 셀이 공백 문자 한 칸이다.
    추가로 pandas를 한 번 거치면 그 None이 float NaN으로 바뀌므로 NaN도 결측 취급한다."""
    if v is None:
        return True
    if isinstance(v, float) and v != v:  # NaN
        return True
    return isinstance(v, str) and not v.strip()


def _num(v):
    if _blank(v):
        return None
    try:
        n = float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None
    return None if n != n else n  # 'nan' 문자열이 float('nan')으로 통과하는 것 차단


def _int(v):
    n = _num(v)
    return int(n) if n is not None else None


def _txt(v):
    return "" if _blank(v) else str(v).strip()


def _status_code(v):
    """'(V) Under construction, more than 50 percent...' -> 'V'"""
    s = _txt(v)
    if s.startswith("(") and ")" in s:
        return s[1:s.index(")")]
    return s


def _ym(y, m):
    """(2027, 6) -> '2027-06'. 연도 없으면 ''."""
    yi, mi = _int(y), _int(m)
    if not yi or yi < 1900 or yi > 2100:
        return ""
    return f"{yi:04d}-{mi:02d}" if mi and 1 <= mi <= 12 else f"{yi:04d}-00"


def read_sheet(wb, sheet: str, colmap: dict) -> pd.DataFrame:
    """헤더명으로 컬럼을 찾아 읽는다 (함정 2: 위치는 시대별로 다름)."""
    if sheet not in wb.sheetnames:
        return pd.DataFrame(columns=list(colmap))
    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)

    header, idx = None, {}
    for r in rows:  # 함정 3: 헤더는 3행이지만 고정하지 않고 탐색
        if r and any(_txt(c) == "Plant ID" for c in r):
            header = [_txt(c) for c in r]
            idx = {std: header.index(xl) for std, xl in colmap.items() if xl in header}
            break
    if header is None:
        return pd.DataFrame(columns=list(colmap))

    out = []
    for r in rows:
        if not r or _blank(r[idx["plant_id"]]):
            continue
        out.append({std: (r[i] if i < len(r) else None) for std, i in idx.items()})
    df = pd.DataFrame(out)
    # 함정 5: 없는 컬럼(MWh/DC 등)은 결측으로 채워 스키마 고정
    for std in colmap:
        if std not in df.columns:
            df[std] = None
    return df[list(colmap)]


def _normalize(df: pd.DataFrame) -> pd.DataFrame:
    for c in df.columns:
        if c in ("plant_id", "entity_id"):
            df[c] = df[c].map(_int)
        elif c in ("mw", "mw_summer", "mwh", "dc_mw", "lat", "lon"):
            df[c] = df[c].map(_num)
        elif c == "status":
            df[c] = df[c].map(_status_code)
        elif c.endswith(("_y", "_m")):
            df[c] = df[c].map(_int)
        else:
            df[c] = df[c].map(_txt)
    return df


def _collapse_ym(df: pd.DataFrame, prefix: str, out: str) -> pd.DataFrame:
    ycol, mcol = f"{prefix}_y", f"{prefix}_m"
    if ycol in df.columns:
        df[out] = [_ym(y, m) for y, m in zip(df[ycol], df[mcol])]
        df = df.drop(columns=[ycol, mcol])
    return df


def parse_planned(path: Path, vintage: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """(빈티지 행, 차원 행) 반환."""
    wb = _open_validated(path)
    try:
        pl = _normalize(read_sheet(wb, "Planned", PLANNED_COLS))
        dim = _normalize(read_sheet(wb, "Planned", DIM_COLS))
    finally:
        wb.close()
    pl = _collapse_ym(pl, "cod", "cod")
    pl.insert(0, "vintage", vintage)
    return pl, dim


def parse_snapshot(path: Path) -> dict[str, pd.DataFrame]:
    """최신 발행월에서 Operating / Retired / Canceled 스냅샷을 뽑는다."""
    wb = _open_validated(path)
    try:
        op = _normalize(read_sheet(wb, "Operating", OPERATING_COLS))
        rt = _normalize(read_sheet(wb, "Retired", RETIRED_COLS))
        cx = _normalize(read_sheet(wb, "Canceled or Postponed", CANCELED_COLS))
    finally:
        wb.close()
    op = _collapse_ym(_collapse_ym(op, "op", "op_ym"), "ret", "ret_ym")
    rt = _collapse_ym(_collapse_ym(rt, "op", "op_ym"), "ret", "ret_ym")
    return {"operating": op, "retired": rt, "canceled": cx}


# ---------------------------------------------------------------- 저장

def _save(df: pd.DataFrame, name: str) -> None:
    EIA_DIR.mkdir(parents=True, exist_ok=True)
    path = EIA_DIR / f"{name}.csv.gz"
    with gzip.open(path, "wt", encoding="utf-8", newline="", compresslevel=9) as f:
        df.to_csv(f, index=False)
    print(f"  saved {path.relative_to(ROOT)}  ({len(df):,} rows, {path.stat().st_size/1e6:.1f} MB)")


def _load(name: str) -> pd.DataFrame | None:
    path = EIA_DIR / f"{name}.csv.gz"
    if not path.exists():
        return None
    return pd.read_csv(path, dtype={"gen_id": str, "cod": str, "vintage": str})


def merge_dim(old: pd.DataFrame | None, new: pd.DataFrame) -> pd.DataFrame:
    """(plant_id, gen_id) 기준 최신값 우선 병합."""
    frames = [new] if old is None else [old, new]
    df = pd.concat(frames, ignore_index=True)
    df = df.drop_duplicates(subset=["plant_id", "gen_id"], keep="last")
    return df.sort_values(["plant_id", "gen_id"]).reset_index(drop=True)


# ---------------------------------------------------------------- 메인

def latest_published(max_lookback: int = 8) -> tuple[int, int] | None:
    """오늘 기준 역순으로 훑어 실제 존재하는 최신 발행월을 찾는다.
    EIA는 데이터월 기준 약 2개월 지연으로 낸다."""
    t = date.today()
    y, m = t.year, t.month
    for _ in range(max_lookback):
        name = f"{MONTHS[m - 1]}_generator{y}.xlsx"
        for sub in ("xls", "archive/xls"):
            req = urllib.request.Request(f"{BASE}/{sub}/{name}", method="HEAD", headers=UA)
            try:
                with urllib.request.urlopen(req, timeout=30) as r:
                    if r.status == 200 and int(r.headers.get("Content-Length") or 0) > STUB_MAX_BYTES:
                        return y, m
            except Exception:
                pass
        m -= 1
        if m < 1:
            y, m = y - 1, 12
    return None


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--backfill", action="store_true",
                    help=f"{BACKFILL_START[0]}-{BACKFILL_START[1]:02d}부터 전부 재수집(로컬 1회)")
    ap.add_argument("--start", default=None, help="백필 시작월 YYYY-MM")
    ap.add_argument("--keep-raw", action="store_true", help="원본 xlsx 캐시 유지(기본: 파싱 후 삭제)")
    ap.add_argument("--workers", type=int, default=4)
    args = ap.parse_args()

    latest = latest_published()
    if latest is None:
        print("!! 최신 발행월을 찾지 못함", file=sys.stderr)
        return 1
    print(f"최신 발행월: {latest[0]}-{latest[1]:02d}")

    vint = _load("vintages")
    have = set(vint["vintage"].unique()) if vint is not None else set()

    if args.backfill:
        start = BACKFILL_START
        if args.start:
            sy, sm = args.start.split("-")
            start = (int(sy), int(sm))
        want = list(_month_iter(start, latest))
    else:
        want = [ym for ym in _month_iter(BACKFILL_START, latest)
                if f"{ym[0]:04d}-{ym[1]:02d}" not in have]
        if vint is None:
            print("!! 빈티지 저장소가 없다. --backfill 로 최초 구축 필요.", file=sys.stderr)
            return 1

    todo = [ym for ym in want if f"{ym[0]:04d}-{ym[1]:02d}" not in have or args.backfill]
    if not todo:
        print("신규 발행월 없음. 변경 없이 종료.")
        return 0
    print(f"수집 대상 {len(todo)}개월: {todo[0][0]}-{todo[0][1]:02d} ~ {todo[-1][0]}-{todo[-1][1]:02d}")

    # 다운로드는 병렬, 파싱은 순차(openpyxl 메모리)
    def dl(ym):
        y, m = ym
        p = fetch_month(y, m)
        print(f"  dl {y}-{m:02d} {'ok' if p else 'FAIL'}", flush=True)
        return ym, p

    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        got = list(ex.map(dl, todo))

    new_vint, dim_acc = [], None
    for (y, m), path in got:
        if path is None:
            print(f"  !! {y}-{m:02d} 다운로드 실패, 건너뜀", file=sys.stderr)
            continue
        vtag = f"{y:04d}-{m:02d}"
        try:
            pl, dim = parse_planned(path, vtag)
        except ValueError as e:
            print(f"  !! {vtag} 파싱 실패: {e}", file=sys.stderr)
            continue
        new_vint.append(pl)
        dim_acc = dim if dim_acc is None else merge_dim(dim_acc, dim)
        print(f"  parsed {vtag}: planned {len(pl):,} rows", flush=True)

    if not new_vint:
        print("!! 파싱된 빈티지 없음", file=sys.stderr)
        return 1

    added = pd.concat(new_vint, ignore_index=True)
    if vint is not None and not args.backfill:
        vint = pd.concat([vint, added], ignore_index=True)
    else:
        vint = added
    vint = vint.drop_duplicates(subset=["vintage", "plant_id", "gen_id"], keep="last")
    vint = vint.sort_values(["vintage", "plant_id", "gen_id"]).reset_index(drop=True)
    _save(vint, "vintages")

    # 최신월 스냅샷(Operating/Retired/Canceled)은 항상 최신본으로 덮어씀
    lp = fetch_month(*latest)
    if lp:
        snap = parse_snapshot(lp)
        for name, df in snap.items():
            _save(df, name)
        dim_acc = merge_dim(dim_acc, snap["operating"][list(DIM_COLS)])

    _save(merge_dim(_load("dim"), dim_acc), "dim")

    if not args.keep_raw and RAW_DIR.exists():
        shutil.rmtree(RAW_DIR, ignore_errors=True)
        print("  raw 캐시 삭제(--keep-raw 로 유지 가능)")

    print(f"완료. 빈티지 {vint['vintage'].nunique()}개월.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
